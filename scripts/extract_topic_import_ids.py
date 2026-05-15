"""Extract topic import IDs from the master category workbook.

Usage:
    python scripts/extract_topic_import_ids.py \
        "/Users/jackcarlson/Downloads/MASTER CATEGORY LIST.xlsx"
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_OUTPUT = Path("config/topic_import_ids_catalog.json")
DEFAULT_SHEET = "SPORTS"
ID_HEADER = "topic_import_id"
LABEL_HEADERS = ("Sport", "League/Category", "topic_name")


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower()


def _header_index(headers: tuple[Any, ...], name: str) -> int:
    normalized = _normalize_header(name)
    for idx, value in enumerate(headers):
        if _normalize_header(value) == normalized:
            return idx
    raise ValueError(f"Missing required column: {name}")


def extract_topic_import_ids(workbook_path: Path, sheet_name: str) -> list[dict[str, str]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Workbook does not contain sheet: {sheet_name}")

        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration as exc:
            raise ValueError(f"Sheet is empty: {sheet_name}") from exc

        id_idx = _header_index(headers, ID_HEADER)
        label_indexes = [
            (header, idx)
            for header in LABEL_HEADERS
            for idx, value in enumerate(headers)
            if _normalize_header(value) == _normalize_header(header)
        ]

        entries_by_id: dict[str, dict[str, str]] = {}
        for row in rows:
            raw_id = row[id_idx] if id_idx < len(row) else None
            topic_id = str(raw_id or "").strip()
            if not topic_id:
                continue

            label_parts = []
            for _, idx in label_indexes:
                raw = row[idx] if idx < len(row) else None
                part = str(raw or "").strip()
                if part and part not in label_parts:
                    label_parts.append(part)
            entries_by_id[topic_id] = {
                "id": topic_id,
                "label": " | ".join(label_parts),
            }

        return [entries_by_id[topic_id] for topic_id in sorted(entries_by_id)]
    finally:
        workbook.close()


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Extract topic_import_id values from the master category workbook."
    )
    parser.add_argument("workbook", type=Path, help="Path to MASTER CATEGORY LIST.xlsx")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="Sheet containing topic_import_id")
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help="Output JSON path relative to the repo root.",
    )
    args = parser.parse_args()

    entries = extract_topic_import_ids(args.workbook, args.sheet)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(entries, indent=2, ensure_ascii=True) + "\n",
        encoding="utf-8",
    )
    print(f"Wrote {len(entries)} topic import IDs to {args.output}")


if __name__ == "__main__":
    main()
